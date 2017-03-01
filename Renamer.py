import logging
from os import listdir
from tkinter import *
from tkinter.filedialog import askdirectory
from tkinter.messagebox import showwarning

from openpyxl import os, load_workbook, Workbook

logger = logging.getLogger('Main')
logger.setLevel(logging.DEBUG)

ch = logging.StreamHandler()
ch.setLevel(logging.INFO)
ch.setFormatter(logging.Formatter('%(name)s - %(levelname)s - %(message)s'))
logger.addHandler(ch)

name_file = 'names.xlsx'
initial_directory = "C:/"


class App:
    def __init__(self, master):
        self.frame = Frame(master)
        self.frame.pack()
        # master.iconbitmap('Icon.ico')
        
        self.folder = ""
        
        self.SelectButton = Button(self.frame, text="Select folder", command=self.scan_folder)
        self.OpenButton = Button(self.frame, text="Open names file", command=self.open_name_file)
        self.Apply = Button(self.frame, text="Apply Names", command=self.apply_names)
        
        self.SelectButton.pack(side=TOP)
        self.OpenButton.pack(side=TOP)
        self.Apply.pack(side=TOP)
    
    def scan_folder(self):
        # Open dialog to ask for desired folder
        self.folder = askdirectory(mustexist=True, initialdir=initial_directory)
        
        logger.debug("Folder: {}".format(self.folder))
        # Read file names in target
        
        try:
            files = listdir(self.folder)
            
            try:
                wb = Workbook()
                ws = wb.active
                ws.append(['File Name', 'New File Name', 'Extension', 'Folder'])
                
                for i in files:
                    logger.info("{}/{}".format(self.folder, i))
                    if os.path.isfile("{}/{}".format(self.folder, i)):
                        filename, file_extension = os.path.splitext(i)
                        logger.debug("{} - {}".format(filename, file_extension))
                        ws.append([filename, '', file_extension, self.folder])
                    
                    if os.path.isdir("{}/{}".format(self.folder, i)):
                        pass
                        # TODO: Scan inner folders for files; Include line for folder and for each file
                
                wb.save(name_file)
            # TODO: Grey out (and lock?) file format column
            except IOError:
                showwarning(title="File inaccessible", message="Please close {}".format(name_file))
            
            self.open_name_file()
        
        except FileNotFoundError:
            pass
        
        logger.debug("{} is closing".format(name_file))
    
    @staticmethod
    def open_name_file():
        logger.debug("open_name_file ran")
        os.startfile(name_file)
    
    @staticmethod
    def apply_names():
        # try:
        wb = load_workbook(name_file, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(row_offset=1):
            if row[1].value == ".":
                # TODO: Delete file
                pass
            elif row[1].value == "":
                pass
            elif row[1].value is None:
                pass
            else:
                os.chdir(row[3].value)
                print("dir: {}".format(os.getcwd()))
                print("in: {}{}".format(row[0].value, row[2].value))
                print("out: {}{}".format(row[1].value, row[2].value))
                # os.rename("{}{}".format(row[0].value, row[2].value),
                #           "{}{}".format(row[1].value, row[2].value))
            
            logger.debug("apply_names ran")


root = Tk()

app = App(root)

root.mainloop()

try:
    root.destroy()
except TclError:
    pass
